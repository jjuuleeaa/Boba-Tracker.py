from flask import Flask, request, jsonify, render_template
import pandas as pd
import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

app = Flask(__name__)

@app.route('/')
def index():
    # Serve the main HTML file
    return render_template('boba_tracker.html')

@app.route('/save_to_excel', methods=['POST'])
def save_to_excel():
    data = request.json.get('data')
    columns = [
        "Date Tried", "Flavor Name", "Flavor Type", "Tea Base",
        "Toppings", "Sweetness Level", "Temperature",
        "Taste Rating", "Notes", "Would Try Again?"
    ]
    df = pd.DataFrame(data, columns=columns)

    # Save the DataFrame to an Excel file with a timestamp
    filename = f"Boba_Flavor_Tracker_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    df.to_excel(filepath, index=False)

    # Load the workbook to apply custom styling
    wb = load_workbook(filepath)
    ws = wb.active

    # Define color styles matching the website
    header_fill = PatternFill(start_color="D4A373", end_color="D4A373", fill_type="solid")  # Header color
    cell_fill = PatternFill(start_color="FBEEA1", end_color="FBEEA1", fill_type="solid")    # Cell background
    header_font = Font(color="FFFFFF", bold=True)  # White, bold text for header
    cell_font = Font(color="4B3832")  # Dark brown text for cell content

    # Apply styles to headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    # Apply styles to data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = cell_fill
            cell.font = cell_font
            cell.alignment = Alignment(wrap_text=True)

    # Save the workbook with styles applied
    wb.save(filepath)

    return jsonify({"message": "Data saved successfully", "filename": filename})

if __name__ == "__main__":
    app.run(debug=True)